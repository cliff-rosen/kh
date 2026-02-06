#!/usr/bin/env python3
"""
Generate the scoring math spreadsheet for Knowledge Horizon.

This script creates an .xlsx file that demonstrates every scoring calculation
actually implemented in the codebase, with working formulas and sample data.

Run: python docs/scoring_math_spreadsheet.py
Output: docs/kh_scoring_math.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =============================================================================
# Styling helpers
# =============================================================================

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SECTION_FONT = Font(bold=True, size=11)
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FORMULA_FONT = Font(italic=True, color="1F4E79")
CODE_FONT = Font(name="Consolas", size=10)
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = BORDER


def style_section_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = SECTION_FONT
        cell.fill = SECTION_FILL
        cell.border = BORDER


def style_data_area(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = BORDER


def auto_width(ws, max_col, min_width=12, max_width=40):
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        best = min_width
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    best = max(best, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[letter].width = best


# =============================================================================
# Tab 1: Constants & Thresholds (reference lookup)
# =============================================================================

def build_tab1_constants(wb):
    ws = wb.create_sheet("1. Constants & Thresholds")

    headers = ["Constant", "Value", "Type", "Source File", "Line(s)", "Purpose"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    style_header_row(ws, 1, len(headers))

    data = [
        ["Semantic Filter Threshold (default)", 0.7, "float 0.0-1.0",
         "schemas/research_stream.py", "112",
         "Minimum LLM score for an article to pass semantic filtering"],
        ["Semantic Filter Score Range", "0.0 - 1.0", "float",
         "pipeline_service.py", "1279-1280",
         "Min/max value passed to AIEvaluationService.score()"],
        ["Relevance Score Range", "0.0 - 10.0", "float",
         "search_providers/base.py", "167",
         "Position-based relevance score range"],
        ["Answer Eval Score Threshold", 0.8, "float 0.0-1.0",
         "iterative_answer_service.py", "30",
         "Minimum LLM score to accept an iterative answer"],
        ["Answer Eval Max Iterations", 3, "int",
         "iterative_answer_service.py", "41",
         "Max generate->evaluate->feedback cycles"],
        ["Deep Research Confidence Threshold", 0.8, "float 0.0-1.0",
         "deep_research_service.py", "143",
         "Below this, first-pass PASS triggers second opinion"],
        ["Deep Research Min Sources", 3, "int",
         "deep_research_service.py", "144",
         "Minimum sources before research can be considered complete"],
        ["Deep Research Max Iterations", 10, "int",
         "deep_research_service.py", "139",
         "Max research loop iterations"],
        ["Deep Research Timeout", 600, "seconds",
         "deep_research_service.py", "140",
         "Hard timeout for the entire research loop"],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_data_area(ws, 2, 1 + len(data), len(headers))

    # Confidence rubric section
    rubric_start = len(data) + 4
    ws.cell(row=rubric_start, column=1, value="Confidence Calibration Rubric (all LLM outputs)")
    ws.cell(row=rubric_start, column=1).font = SECTION_FONT
    ws.merge_cells(start_row=rubric_start, start_column=1, end_row=rubric_start, end_column=4)

    rubric = [
        ["Score Range", "Meaning", "Used In"],
        ["0.9 - 1.0", "Explicit statement in source text", "filter, score, extract, extract_fields"],
        ["0.7 - 0.89", "Strong inference with clear supporting context", "filter, score, extract, extract_fields"],
        ["0.4 - 0.69", "Weak inference or ambiguous evidence", "filter, score, extract, extract_fields"],
        ["Below 0.4", "Insufficient evidence", "filter, score, extract, extract_fields"],
    ]
    for i, row_data in enumerate(rubric):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=rubric_start + 1 + i, column=c, value=val)
    style_header_row(ws, rubric_start + 1, 3)
    style_data_area(ws, rubric_start + 2, rubric_start + len(rubric), 3)

    auto_width(ws, len(headers))


# =============================================================================
# Tab 2: Search Relevance Scoring
# =============================================================================

def build_tab2_relevance(wb):
    ws = wb.create_sheet("2. Search Relevance Score")

    # Formula explanation
    ws.cell(row=1, column=1, value="Formula: score = 10.0 × (1.0 - (position - 1) / min(total_results, 100))")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.merge_cells("A1:G1")

    ws.cell(row=2, column=1, value="Source: backend/services/search_providers/base.py:152-168")
    ws.cell(row=2, column=1).font = CODE_FONT
    ws.cell(row=3, column=1, value="Clamp: max(0.0, min(10.0, result))  •  If total_results <= 0, return 5.0")
    ws.cell(row=3, column=1).font = CODE_FONT

    # Main table
    headers = ["Article", "Position", "Total Results", "min(total, 100)",
               "Decay Fraction", "Raw Score", "Clamped Score"]
    row = 5
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))

    # Sample data: 10 articles from a search with 50 results
    articles = [
        ("CRISPR gene therapy in sickle cell", 1, 50),
        ("Novel CAR-T approach for leukemia", 2, 50),
        ("mRNA vaccine platform review", 5, 50),
        ("Insulin resistance biomarkers", 10, 50),
        ("Cardiology outcomes meta-analysis", 25, 50),
        ("Alzheimer's tau protein study", 50, 50),
        ("Rare disease case report", 1, 5),
        ("Phase III trial results", 3, 5),
        ("Epidemiology survey data", 5, 5),
        ("Single result scenario", 1, 1),
    ]

    for i, (title, pos, total) in enumerate(articles):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=title)
        ws.cell(row=r, column=2, value=pos)
        ws.cell(row=r, column=3, value=total)
        # D: =MIN(C, 100)
        ws.cell(row=r, column=4).value = f"=MIN(C{r},100)"
        ws.cell(row=r, column=4).font = FORMULA_FONT
        # E: =(B-1)/D  (decay fraction)
        ws.cell(row=r, column=5).value = f"=(B{r}-1)/D{r}"
        ws.cell(row=r, column=5).font = FORMULA_FONT
        ws.cell(row=r, column=5).number_format = "0.000"
        # F: =10*(1-E)
        ws.cell(row=r, column=6).value = f"=10*(1-E{r})"
        ws.cell(row=r, column=6).font = FORMULA_FONT
        ws.cell(row=r, column=6).number_format = "0.00"
        # G: =MAX(0, MIN(10, F))
        ws.cell(row=r, column=7).value = f"=MAX(0,MIN(10,F{r}))"
        ws.cell(row=r, column=7).font = FORMULA_FONT
        ws.cell(row=r, column=7).number_format = "0.00"

    end_row = row + len(articles)
    style_data_area(ws, row + 1, end_row, len(headers))

    # Edge case note
    note_row = end_row + 2
    ws.cell(row=note_row, column=1, value="Edge case: if total_results <= 0 → returns 5.0 (no formula, hardcoded)")
    ws.cell(row=note_row, column=1).font = Font(italic=True)

    auto_width(ws, len(headers))


# =============================================================================
# Tab 3: Semantic Filter Scoring
# =============================================================================

def build_tab3_semantic_filter(wb):
    ws = wb.create_sheet("3. Semantic Filter")

    ws.cell(row=1, column=1, value="Semantic Filter: LLM scores each article 0.0-1.0, then threshold comparison")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.merge_cells("A1:I1")
    ws.cell(row=2, column=1, value="Source: pipeline_service.py:1220-1296, wip_article_service.py:295-335")
    ws.cell(row=2, column=1).font = CODE_FONT
    ws.cell(row=3, column=1, value="Decision: passed = (filter_score >= threshold)")
    ws.cell(row=3, column=1).font = CODE_FONT

    # Config section
    ws.cell(row=5, column=1, value="Configuration")
    style_section_row(ws, 5, 3)
    ws.cell(row=6, column=1, value="Threshold")
    ws.cell(row=6, column=2, value=0.7)
    ws.cell(row=6, column=2).number_format = "0.0"
    ws.cell(row=6, column=3, value="← change this to see pass/fail recalculate")
    ws.cell(row=6, column=3).font = Font(italic=True, color="808080")

    # Article table
    row = 8
    headers = ["Article ID", "Title", "Abstract Snippet", "Journal",
               "LLM Score (0-1)", "LLM Confidence", "Threshold",
               "Passed?", "Reasoning"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))

    articles = [
        ("ART-001", "CRISPR gene therapy in sickle cell disease",
         "We report results of a phase I trial...", "NEJM",
         0.92, 0.95, "High relevance to gene therapy criteria"),
        ("ART-002", "Novel CAR-T approach for leukemia",
         "Chimeric antigen receptor T-cell therapy...", "Blood",
         0.85, 0.88, "Strong match to immunotherapy search"),
        ("ART-003", "mRNA vaccine platform review",
         "A comprehensive review of mRNA technology...", "Nature Reviews",
         0.71, 0.72, "Borderline - tangentially related"),
        ("ART-004", "Weather patterns in the Sahara",
         "Analysis of rainfall patterns shows...", "J Climatology",
         0.12, 0.91, "Not relevant to biomedical query"),
        ("ART-005", "Insulin resistance biomarkers",
         "Novel biomarkers for predicting type 2...", "Diabetes Care",
         0.68, 0.80, "Below threshold - weak connection to primary topic"),
        ("ART-006", "Cardiology outcomes meta-analysis",
         "This meta-analysis of 42 RCTs...", "JAMA Cardiology",
         0.45, 0.85, "Different therapeutic area"),
        ("ART-007", "Rare disease gene panel study",
         "Whole exome sequencing in 200 patients...", "Genetics in Medicine",
         0.78, 0.70, "Moderate match - genetic approach relevant"),
        ("ART-008", "Cooking recipes for healthy living",
         "A guide to Mediterranean diet...", "Food Science",
         0.05, 0.97, "Completely off-topic"),
    ]

    for i, (aid, title, abstract, journal, score, confidence, reasoning) in enumerate(articles):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=aid)
        ws.cell(row=r, column=2, value=title)
        ws.cell(row=r, column=3, value=abstract)
        ws.cell(row=r, column=4, value=journal)
        ws.cell(row=r, column=5, value=score)
        ws.cell(row=r, column=5).number_format = "0.00"
        ws.cell(row=r, column=6, value=confidence)
        ws.cell(row=r, column=6).number_format = "0.00"
        # G: threshold reference (linked to config cell B6)
        ws.cell(row=r, column=7).value = "=$B$6"
        ws.cell(row=r, column=7).font = FORMULA_FONT
        ws.cell(row=r, column=7).number_format = "0.0"
        # H: passed? = score >= threshold
        ws.cell(row=r, column=8).value = f'=IF(E{r}>=G{r},"PASS","FAIL")'
        ws.cell(row=r, column=8).font = FORMULA_FONT
        ws.cell(row=r, column=9, value=reasoning)

    end_data = row + len(articles)
    style_data_area(ws, row + 1, end_data, len(headers))

    # Apply conditional formatting manually for pass/fail
    for i in range(len(articles)):
        r = row + 1 + i
        score = articles[i][4]
        cell = ws.cell(row=r, column=8)
        if score >= 0.7:
            cell.fill = PASS_FILL
        else:
            cell.fill = FAIL_FILL

    # Summary section
    summary_row = end_data + 2
    ws.cell(row=summary_row, column=1, value="Summary")
    style_section_row(ws, summary_row, 4)
    sr = summary_row
    ws.cell(row=sr + 1, column=1, value="Total Articles")
    ws.cell(row=sr + 1, column=2).value = f"=COUNTA(A{row+1}:A{end_data})"
    ws.cell(row=sr + 1, column=2).font = FORMULA_FONT
    ws.cell(row=sr + 2, column=1, value="Passed")
    ws.cell(row=sr + 2, column=2).value = f'=COUNTIF(H{row+1}:H{end_data},"PASS")'
    ws.cell(row=sr + 2, column=2).font = FORMULA_FONT
    ws.cell(row=sr + 3, column=1, value="Rejected")
    ws.cell(row=sr + 3, column=2).value = f'=COUNTIF(H{row+1}:H{end_data},"FAIL")'
    ws.cell(row=sr + 3, column=2).font = FORMULA_FONT
    ws.cell(row=sr + 4, column=1, value="Pass Rate")
    ws.cell(row=sr + 4, column=2).value = f"=B{sr+2}/B{sr+1}"
    ws.cell(row=sr + 4, column=2).font = FORMULA_FONT
    ws.cell(row=sr + 4, column=2).number_format = "0.0%"

    # Prompt template section
    prompt_row = sr + 6
    ws.cell(row=prompt_row, column=1, value="LLM Prompt Template (from pipeline_service.py:1260-1270)")
    ws.cell(row=prompt_row, column=1).font = SECTION_FONT
    ws.merge_cells(start_row=prompt_row, start_column=1, end_row=prompt_row, end_column=6)
    prompt_lines = [
        "## Article",
        "Title: {title}",
        "Abstract: {abstract}",
        "AI Summary: {summary}",
        "Journal: {journal}",
        "Authors: {authors}",
        "",
        "## Task",
        "{filter_criteria}",
        "",
        "Score from {min_value} to {max_value}.",
    ]
    for j, line in enumerate(prompt_lines):
        ws.cell(row=prompt_row + 1 + j, column=1, value=line)
        ws.cell(row=prompt_row + 1 + j, column=1).font = CODE_FONT

    auto_width(ws, len(headers))


# =============================================================================
# Tab 4: Iterative Answer Evaluation
# =============================================================================

def build_tab4_iterative_answer(wb):
    ws = wb.create_sheet("4. Iterative Answer Eval")

    ws.cell(row=1, column=1, value="Iterative Answer Generation: generate → evaluate → feedback → repeat")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.merge_cells("A1:H1")
    ws.cell(row=2, column=1, value="Source: iterative_answer_service.py:27-183")
    ws.cell(row=2, column=1).font = CODE_FONT
    ws.cell(row=3, column=1, value="Decision: if eval_score >= threshold → ACCEPT, else → iterate with feedback (up to max_iterations)")
    ws.cell(row=3, column=1).font = CODE_FONT

    # Config
    ws.cell(row=5, column=1, value="Configuration")
    style_section_row(ws, 5, 4)
    ws.cell(row=6, column=1, value="Score Threshold")
    ws.cell(row=6, column=2, value=0.8)
    ws.cell(row=6, column=2).number_format = "0.0"
    ws.cell(row=7, column=1, value="Max Iterations")
    ws.cell(row=7, column=2, value=3)
    ws.cell(row=8, column=1, value="Score Range")
    ws.cell(row=8, column=2, value="0.0 - 1.0 (LLM-assigned)")

    # Example iteration trace
    row = 10
    headers = ["Iteration", "Phase", "LLM Score (0-1)", "Threshold",
               "Meets Threshold?", "Action", "Feedback Summary"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))

    iterations = [
        (1, "Generate + Evaluate", 0.55, "Answer lacks specificity; missing citations"),
        (2, "Generate + Evaluate", 0.72, "Improved detail but weak conclusion"),
        (3, "Generate + Evaluate", 0.84, "Meets criteria"),
    ]

    for i, (iter_num, phase, score, feedback) in enumerate(iterations):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=iter_num)
        ws.cell(row=r, column=2, value=phase)
        ws.cell(row=r, column=3, value=score)
        ws.cell(row=r, column=3).number_format = "0.00"
        ws.cell(row=r, column=4).value = "=$B$6"
        ws.cell(row=r, column=4).font = FORMULA_FONT
        ws.cell(row=r, column=4).number_format = "0.0"
        # Meets threshold?
        ws.cell(row=r, column=5).value = f'=IF(C{r}>=$B$6,"YES","NO")'
        ws.cell(row=r, column=5).font = FORMULA_FONT
        # Action
        ws.cell(row=r, column=6).value = f'=IF(C{r}>=$B$6,"ACCEPT",IF(A{r}<$B$7,"RETRY with feedback","RETURN BEST"))'
        ws.cell(row=r, column=6).font = FORMULA_FONT
        ws.cell(row=r, column=7, value=feedback)

    end_data = row + len(iterations)
    style_data_area(ws, row + 1, end_data, len(headers))

    # Outcome section
    out_row = end_data + 2
    ws.cell(row=out_row, column=1, value="Outcome")
    style_section_row(ws, out_row, 4)
    ws.cell(row=out_row + 1, column=1, value="Best Score")
    ws.cell(row=out_row + 1, column=2).value = f"=MAX(C{row+1}:C{end_data})"
    ws.cell(row=out_row + 1, column=2).font = FORMULA_FONT
    ws.cell(row=out_row + 1, column=2).number_format = "0.00"
    ws.cell(row=out_row + 2, column=1, value="Success?")
    ws.cell(row=out_row + 2, column=2).value = f'=IF(B{out_row+1}>=$B$6,"YES","NO")'
    ws.cell(row=out_row + 2, column=2).font = FORMULA_FONT
    ws.cell(row=out_row + 3, column=1, value="Total Iterations Used")
    ws.cell(row=out_row + 3, column=2).value = f"=COUNTA(A{row+1}:A{end_data})"
    ws.cell(row=out_row + 3, column=2).font = FORMULA_FONT

    # Second example: failure case
    fail_row = out_row + 6
    ws.cell(row=fail_row, column=1, value="Example 2: Max iterations reached (failure case)")
    ws.cell(row=fail_row, column=1).font = SECTION_FONT
    ws.merge_cells(start_row=fail_row, start_column=1, end_row=fail_row, end_column=6)

    headers2 = ["Iteration", "Phase", "LLM Score", "Threshold", "Meets?", "Action", "Feedback"]
    fh_row = fail_row + 1
    for c, h in enumerate(headers2, 1):
        ws.cell(row=fh_row, column=c, value=h)
    style_header_row(ws, fh_row, len(headers2))

    fail_iters = [
        (1, "Generate + Evaluate", 0.35, "Completely off-topic response"),
        (2, "Generate + Evaluate", 0.52, "Better but still lacks depth"),
        (3, "Generate + Evaluate", 0.61, "Improved but below 0.8 threshold"),
    ]
    for i, (it, phase, score, fb) in enumerate(fail_iters):
        r = fh_row + 1 + i
        ws.cell(row=r, column=1, value=it)
        ws.cell(row=r, column=2, value=phase)
        ws.cell(row=r, column=3, value=score)
        ws.cell(row=r, column=3).number_format = "0.00"
        ws.cell(row=r, column=4).value = "=$B$6"
        ws.cell(row=r, column=4).font = FORMULA_FONT
        ws.cell(row=r, column=5).value = f'=IF(C{r}>=$B$6,"YES","NO")'
        ws.cell(row=r, column=5).font = FORMULA_FONT
        ws.cell(row=r, column=6).value = f'=IF(C{r}>=$B$6,"ACCEPT",IF(A{r}<$B$7,"RETRY","RETURN BEST"))'
        ws.cell(row=r, column=6).font = FORMULA_FONT
        ws.cell(row=r, column=7, value=fb)

    fail_end = fh_row + len(fail_iters)
    style_data_area(ws, fh_row + 1, fail_end, len(headers2))

    fo_row = fail_end + 1
    ws.cell(row=fo_row, column=1, value="Outcome: success=False, returns best answer (iteration with max score)")
    ws.cell(row=fo_row, column=1).font = Font(italic=True, color="C00000")
    ws.cell(row=fo_row + 1, column=1, value="Best Score")
    ws.cell(row=fo_row + 1, column=2).value = f"=MAX(C{fh_row+1}:C{fail_end})"
    ws.cell(row=fo_row + 1, column=2).font = FORMULA_FONT
    ws.cell(row=fo_row + 1, column=2).number_format = "0.00"

    auto_width(ws, len(headers))


# =============================================================================
# Tab 5: Deep Research Evaluation
# =============================================================================

def build_tab5_deep_research(wb):
    ws = wb.create_sheet("5. Deep Research Eval")

    ws.cell(row=1, column=1, value="Deep Research: Two-tier evaluation with second opinion on low-confidence passes")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.merge_cells("A1:I1")
    ws.cell(row=2, column=1, value="Source: deep_research_service.py:439-572")
    ws.cell(row=2, column=1).font = CODE_FONT

    # Config
    ws.cell(row=4, column=1, value="Configuration")
    style_section_row(ws, 4, 4)
    ws.cell(row=5, column=1, value="Confidence Threshold")
    ws.cell(row=5, column=2, value=0.8)
    ws.cell(row=5, column=2).number_format = "0.0"
    ws.cell(row=5, column=3, value="← Below this, PASS triggers second opinion")
    ws.cell(row=5, column=3).font = Font(italic=True, color="808080")
    ws.cell(row=6, column=1, value="Min Sources")
    ws.cell(row=6, column=2, value=3)
    ws.cell(row=7, column=1, value="Max Iterations")
    ws.cell(row=7, column=2, value=10)
    ws.cell(row=8, column=1, value="Timeout")
    ws.cell(row=8, column=2, value="600s (10 min)")

    # Decision tree
    dt_row = 10
    ws.cell(row=dt_row, column=1, value="Evaluation Decision Tree")
    style_section_row(ws, dt_row, 6)
    tree = [
        ["First Evaluator Result", "Confidence", "Path", "Next Action"],
        ["FAIL", "any", "→ Continue Loop", "Return gaps, search for more info"],
        ["PASS", "< threshold (e.g. 0.65)", "→ Second Opinion", "Another LLM reviews the assessment"],
        ["PASS", ">= threshold (e.g. 0.90)", "→ Exit Loop", "Research is sufficient, synthesize answer"],
    ]
    for i, row_data in enumerate(tree):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=dt_row + 1 + i, column=c, value=val)
    style_header_row(ws, dt_row + 1, 4)
    style_data_area(ws, dt_row + 2, dt_row + len(tree), 4)

    # Second opinion outcomes
    so_row = dt_row + len(tree) + 2
    ws.cell(row=so_row, column=1, value="Second Opinion Outcomes")
    style_section_row(ws, so_row, 6)
    so_tree = [
        ["Second Opinion Result", "Action"],
        ["confirmed = True", "Accept research as sufficient, exit loop"],
        ["confirmed = False", "Replace evaluation with second opinion's gaps, continue loop"],
    ]
    for i, row_data in enumerate(so_tree):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=so_row + 1 + i, column=c, value=val)
    style_header_row(ws, so_row + 1, 2)
    style_data_area(ws, so_row + 2, so_row + len(so_tree), 2)

    # Example trace
    trace_row = so_row + len(so_tree) + 2
    ws.cell(row=trace_row, column=1, value="Example Research Trace")
    style_section_row(ws, trace_row, 9)

    t_headers = ["Iteration", "Action", "Sources Found", "Total Sources",
                 "First Eval: Passed?", "First Eval: Confidence",
                 "Needs 2nd Opinion?", "2nd Opinion Result", "Loop Decision"]
    th_row = trace_row + 1
    for c, h in enumerate(t_headers, 1):
        ws.cell(row=th_row, column=c, value=h)
    style_header_row(ws, th_row, len(t_headers))

    trace_data = [
        (1, "PubMed search: 'CRISPR sickle cell'", 4, 4,
         "FAIL", 0.3, "N/A", "N/A", "Continue - 3 gaps found"),
        (2, "Web search: 'FDA CRISPR approval 2024'", 3, 7,
         "FAIL", 0.55, "N/A", "N/A", "Continue - 1 gap remaining"),
        (3, "PubMed search: 'Casgevy outcomes'", 2, 9,
         "PASS", 0.65, "YES (0.65 < 0.8)", "confirmed=True, conf=0.78",
         "EXIT - 2nd opinion confirmed"),
    ]

    for i, row_data in enumerate(trace_data):
        r = th_row + 1 + i
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
            if isinstance(val, float):
                ws.cell(row=r, column=c).number_format = "0.00"

    trace_end = th_row + len(trace_data)
    style_data_area(ws, th_row + 1, trace_end, len(t_headers))

    # Alternative trace: second opinion disagrees
    alt_row = trace_end + 2
    ws.cell(row=alt_row, column=1, value="Alternative: Second opinion finds more gaps")
    ws.cell(row=alt_row, column=1).font = SECTION_FONT
    ws.merge_cells(start_row=alt_row, start_column=1, end_row=alt_row, end_column=6)

    ah_row = alt_row + 1
    for c, h in enumerate(t_headers, 1):
        ws.cell(row=ah_row, column=c, value=h)
    style_header_row(ws, ah_row, len(t_headers))

    alt_data = [
        (1, "PubMed search", 5, 5,
         "FAIL", 0.4, "N/A", "N/A", "Continue"),
        (2, "Web search", 3, 8,
         "PASS", 0.62, "YES (0.62 < 0.8)", "confirmed=False, 2 more gaps",
         "CONTINUE - evaluation replaced with 2nd opinion gaps"),
        (3, "PubMed search (targeted)", 2, 10,
         "PASS", 0.88, "NO (0.88 >= 0.8)", "N/A",
         "EXIT - high confidence pass"),
    ]

    for i, row_data in enumerate(alt_data):
        r = ah_row + 1 + i
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
            if isinstance(val, float):
                ws.cell(row=r, column=c).number_format = "0.00"

    style_data_area(ws, ah_row + 1, ah_row + len(alt_data), len(t_headers))

    auto_width(ws, len(t_headers))


# =============================================================================
# Tab 6: Full Pipeline Flow
# =============================================================================

def build_tab6_pipeline(wb):
    ws = wb.create_sheet("6. Full Pipeline Flow")

    ws.cell(row=1, column=1, value="End-to-End Pipeline: How scores flow from search to report")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.merge_cells("A1:I1")
    ws.cell(row=2, column=1, value="Source: pipeline_service.py (orchestrates all stages)")
    ws.cell(row=2, column=1).font = CODE_FONT

    # Pipeline stages
    row = 4
    ws.cell(row=row, column=1, value="Pipeline Stages (in execution order)")
    style_section_row(ws, row, 8)

    stage_headers = ["Stage", "Input", "Scoring Method", "Output", "Stored In", "Decision Logic"]
    sh_row = row + 1
    for c, h in enumerate(stage_headers, 1):
        ws.cell(row=sh_row, column=c, value=h)
    style_header_row(ws, sh_row, len(stage_headers))

    stages = [
        ["1. Article Retrieval",
         "Search queries → search providers",
         "Position-based: 10×(1-(pos-1)/min(total,100))",
         "relevance_score (0-10)",
         "WipArticle (initially)",
         "All results included"],
        ["2. Deduplication",
         "WipArticles from retrieval",
         "None (exact match on DOI/title)",
         "is_duplicate flag",
         "WipArticle.is_duplicate",
         "Duplicates excluded from further stages"],
        ["3. Semantic Filtering",
         "Non-duplicate WipArticles",
         "LLM score() → 0.0-1.0",
         "filter_score + reasoning",
         "WipArticle.filter_score",
         "passed = (score >= threshold, default 0.7)"],
        ["4. Report Creation",
         "Passed articles",
         "None (inclusion only)",
         "ReportArticleAssociation",
         "report_article_association table",
         "filter_score → relevance_score"],
        ["5. Categorization",
         "Report articles",
         "LLM classification",
         "Category assignments",
         "ReportArticleAssociation.categories",
         "Articles assigned to stream categories"],
        ["6. Enrichment",
         "Report + categorized articles",
         "LLM summarization",
         "Executive summary, category summaries",
         "Report.enrichments (JSON)",
         "N/A - all articles enriched"],
    ]

    for i, row_data in enumerate(stages):
        r = sh_row + 1 + i
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    stage_end = sh_row + len(stages)
    style_data_area(ws, sh_row + 1, stage_end, len(stage_headers))

    # Worked example
    ex_row = stage_end + 2
    ws.cell(row=ex_row, column=1, value="Worked Example: 6 articles through the pipeline")
    style_section_row(ws, ex_row, 9)

    ex_headers = ["Article", "Search Position", "Total Results",
                  "Relevance Score (0-10)", "Is Duplicate?",
                  "Filter Score (0-1)", "Filter Threshold",
                  "Passed Filter?", "In Report?"]
    eh_row = ex_row + 1
    for c, h in enumerate(ex_headers, 1):
        ws.cell(row=eh_row, column=c, value=h)
    style_header_row(ws, eh_row, len(ex_headers))

    # Config cell for threshold
    ws.cell(row=ex_row, column=8, value="Threshold:")
    ws.cell(row=ex_row, column=9, value=0.7)
    ws.cell(row=ex_row, column=9).number_format = "0.0"

    examples = [
        ("CRISPR sickle cell trial", 1, 50, False, 0.92),
        ("CAR-T leukemia approach", 2, 50, False, 0.85),
        ("CRISPR sickle cell trial (dup)", 3, 50, True, None),
        ("mRNA vaccine review", 5, 50, False, 0.71),
        ("Sahara weather patterns", 10, 50, False, 0.12),
        ("Insulin resistance study", 25, 50, False, 0.68),
    ]

    for i, (title, pos, total, is_dup, filter_score) in enumerate(examples):
        r = eh_row + 1 + i
        ws.cell(row=r, column=1, value=title)
        ws.cell(row=r, column=2, value=pos)
        ws.cell(row=r, column=3, value=total)
        # D: relevance score formula
        ws.cell(row=r, column=4).value = f"=MAX(0,MIN(10,10*(1-(B{r}-1)/MIN(C{r},100))))"
        ws.cell(row=r, column=4).font = FORMULA_FONT
        ws.cell(row=r, column=4).number_format = "0.00"
        # E: is duplicate
        ws.cell(row=r, column=5, value="YES" if is_dup else "NO")
        if is_dup:
            ws.cell(row=r, column=5).fill = FAIL_FILL
        # F: filter score (blank if duplicate)
        if filter_score is not None:
            ws.cell(row=r, column=6, value=filter_score)
            ws.cell(row=r, column=6).number_format = "0.00"
        else:
            ws.cell(row=r, column=6, value="(skipped)")
            ws.cell(row=r, column=6).font = Font(italic=True, color="808080")
        # G: threshold
        ws.cell(row=r, column=7).value = f"=$I${ex_row}"
        ws.cell(row=r, column=7).font = FORMULA_FONT
        ws.cell(row=r, column=7).number_format = "0.0"
        # H: passed filter?
        if is_dup:
            ws.cell(row=r, column=8, value="N/A (duplicate)")
            ws.cell(row=r, column=8).font = Font(italic=True, color="808080")
        else:
            ws.cell(row=r, column=8).value = f'=IF(E{r}="YES","N/A",IF(F{r}>=$I${ex_row},"PASS","FAIL"))'
            ws.cell(row=r, column=8).font = FORMULA_FONT
        # I: in report?
        if is_dup:
            ws.cell(row=r, column=9, value="NO")
            ws.cell(row=r, column=9).fill = FAIL_FILL
        elif filter_score is not None and filter_score >= 0.7:
            ws.cell(row=r, column=9).value = f'=IF(AND(E{r}="NO",H{r}="PASS"),"YES","NO")'
            ws.cell(row=r, column=9).font = FORMULA_FONT
        else:
            ws.cell(row=r, column=9).value = f'=IF(AND(E{r}="NO",H{r}="PASS"),"YES","NO")'
            ws.cell(row=r, column=9).font = FORMULA_FONT

    ex_end = eh_row + len(examples)
    style_data_area(ws, eh_row + 1, ex_end, len(ex_headers))

    # Summary
    sum_row = ex_end + 2
    ws.cell(row=sum_row, column=1, value="Pipeline Summary")
    style_section_row(ws, sum_row, 4)
    ws.cell(row=sum_row + 1, column=1, value="Retrieved")
    ws.cell(row=sum_row + 1, column=2, value=len(examples))
    ws.cell(row=sum_row + 2, column=1, value="Deduplicated (removed)")
    ws.cell(row=sum_row + 2, column=2).value = f'=COUNTIF(E{eh_row+1}:E{ex_end},"YES")'
    ws.cell(row=sum_row + 2, column=2).font = FORMULA_FONT
    ws.cell(row=sum_row + 3, column=1, value="Passed Filter")
    ws.cell(row=sum_row + 3, column=2).value = f'=COUNTIF(H{eh_row+1}:H{ex_end},"PASS")'
    ws.cell(row=sum_row + 3, column=2).font = FORMULA_FONT
    ws.cell(row=sum_row + 4, column=1, value="In Final Report")
    ws.cell(row=sum_row + 4, column=2).value = f'=COUNTIF(I{eh_row+1}:I{ex_end},"YES")'
    ws.cell(row=sum_row + 4, column=2).font = FORMULA_FONT

    auto_width(ws, len(ex_headers))


# =============================================================================
# Tab 7: AIEvaluationService API
# =============================================================================

def build_tab7_eval_api(wb):
    ws = wb.create_sheet("7. AIEvaluationService API")

    ws.cell(row=1, column=1, value="AIEvaluationService: Unified LLM evaluation interface")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.merge_cells("A1:G1")
    ws.cell(row=2, column=1, value="Source: backend/services/ai_evaluation_service.py")
    ws.cell(row=2, column=1).font = CODE_FONT

    row = 4
    headers = ["Operation", "Input", "Output Type", "Output Schema",
               "Score Range", "Used For"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))

    ops = [
        ["filter()",
         "items + prompt_template",
         "Boolean",
         "{value: bool, confidence: 0-1, reasoning?: str}",
         "N/A (true/false)",
         "Yes/No classification of articles"],
        ["score()",
         "items + prompt_template + min/max",
         "Float",
         "{value: float, confidence: 0-1, reasoning?: str}",
         "Configurable (default 0-1)",
         "Semantic filtering, relevance rating"],
        ["extract()",
         "items + prompt_template + output_type",
         "Any (text/number/bool/enum)",
         "{value: any|null, confidence: 0-1, reasoning?: str}",
         "N/A",
         "Single value extraction from text"],
        ["extract_fields()",
         "items + prompt_template + schema",
         "Structured (JSON)",
         "{fields: {schema}, confidence: 0-1, reasoning?: str}",
         "N/A",
         "Multi-field extraction per schema"],
    ]

    for i, row_data in enumerate(ops):
        r = row + 1 + i
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_data_area(ws, row + 1, row + len(ops), len(headers))

    # All outputs section
    ao_row = row + len(ops) + 2
    ws.cell(row=ao_row, column=1, value="Common Output Fields (all operations)")
    style_section_row(ws, ao_row, 5)

    ao_headers = ["Field", "Type", "Range", "Required?", "Description"]
    aoh_row = ao_row + 1
    for c, h in enumerate(ao_headers, 1):
        ws.cell(row=aoh_row, column=c, value=h)
    style_header_row(ws, aoh_row, len(ao_headers))

    fields = [
        ["value", "varies by operation", "varies", "Yes", "The primary result (bool, float, string, object)"],
        ["confidence", "float", "0.0 - 1.0", "Yes", "LLM confidence in its answer based on evidence quality"],
        ["reasoning", "string", "N/A", "Optional", "Brief explanation (when include_reasoning=True)"],
    ]

    for i, row_data in enumerate(fields):
        r = aoh_row + 1 + i
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_data_area(ws, aoh_row + 1, aoh_row + len(fields), len(ao_headers))

    auto_width(ws, len(headers))


# =============================================================================
# Tab 8: What's NOT Implemented
# =============================================================================

def build_tab8_not_implemented(wb):
    ws = wb.create_sheet("8. NOT Implemented")

    ws.cell(row=1, column=1, value="Features from original spec that DO NOT EXIST in the codebase")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12, color="C00000")
    ws.merge_cells("A1:E1")
    ws.cell(row=2, column=1, value="The original spreadsheet spec described a hypothetical scoring pipeline. This tab clarifies what is real vs fictional.")
    ws.cell(row=2, column=1).font = Font(italic=True)
    ws.merge_cells("A2:E2")

    row = 4
    headers = ["Fictional Feature", "What It Described", "What Actually Exists", "Gap"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    style_header_row(ws, row, len(headers))

    gaps = [
        ["Pattern Library (509 patterns)",
         "Predefined patterns (PAT-001, PAT-002) with confidence %, keywords",
         "No pattern library. Semantic filter uses freeform LLM criteria, not pattern matching",
         "Major - would need new subsystem"],
        ["Signal Intake",
         "Raw intelligence items from news, social, court filings scored against patterns",
         "Articles come from search providers (PubMed, web). No 'signal' concept",
         "Major - different data model"],
        ["Pattern Matching Matrix",
         "Each signal scored against each pattern → matrix of confidence %",
         "Single LLM score per article against one criteria string",
         "Major - 1D scoring vs N×M matrix"],
        ["Lens Scoring (pattern/legal/social/news/anomaly)",
         "5 weighted lens scores aggregated to initial confidence",
         "No lens system. Single filter_score (0-1) per article",
         "Major - no multi-dimensional scoring"],
        ["Lens Weights (35/20/20/15/10%)",
         "Pattern=35%, Legal=20%, Social=20%, News=15%, Anomaly=10%",
         "No weighted aggregation. Threshold is binary pass/fail",
         "Major - no weighted scoring"],
        ["Evidence Tiers (1-4) with point caps",
         "Court=5pts/15cap, News=2pts/8cap, Pattern=1pt/5cap, Social=0.3pts/3cap",
         "No evidence tier system. All articles treated equally after filter",
         "Major - no evidence accumulation"],
        ["Evidence Bonus",
         "Evidence points add bonus % to initial confidence",
         "No confidence adjustment mechanism",
         "Major - no post-hoc adjustment"],
        ["Validated Confidence",
         "Initial confidence + evidence bonus = final score",
         "filter_score IS the final score. No multi-stage confidence",
         "Major - single-stage only"],
        ["Hypothesis Generation",
         "Signals roll up into a named hypothesis with entity/domain",
         "Reports contain articles with summaries. No 'hypothesis' abstraction",
         "Moderate - different conceptual model"],
        ["Decision Tab (pursue/refine/reject)",
         "Lawyer fills in strengths/weaknesses/next steps with dropdown",
         "No decision workflow. Reports are the final output",
         "Moderate - UI/workflow gap"],
    ]

    for i, row_data in enumerate(gaps):
        r = row + 1 + i
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
            ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True)

    style_data_area(ws, row + 1, row + len(gaps), len(headers))

    # What DOES exist
    exist_row = row + len(gaps) + 2
    ws.cell(row=exist_row, column=1, value="What IS Implemented (scoring-related)")
    style_section_row(ws, exist_row, 4)

    exist_headers = ["Feature", "Description", "Key File"]
    exh_row = exist_row + 1
    for c, h in enumerate(exist_headers, 1):
        ws.cell(row=exh_row, column=c, value=h)
    style_header_row(ws, exh_row, len(exist_headers))

    exists = [
        ["Position-based relevance", "Linear decay 0-10 based on search result position", "search_providers/base.py"],
        ["LLM semantic filtering", "Score 0-1 with configurable threshold (default 0.7)", "pipeline_service.py"],
        ["Iterative answer eval", "Generate→evaluate→feedback loop, threshold 0.8, max 3 iter", "iterative_answer_service.py"],
        ["Deep research eval", "Two-tier: first eval + second opinion if confidence < 0.8", "deep_research_service.py"],
        ["Confidence calibration rubric", "Consistent 0-1 confidence scale across all LLM operations", "ai_evaluation_service.py"],
        ["Unified eval API", "filter(), score(), extract(), extract_fields() with batch support", "ai_evaluation_service.py"],
    ]

    for i, row_data in enumerate(exists):
        r = exh_row + 1 + i
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    style_data_area(ws, exh_row + 1, exh_row + len(exists), len(exist_headers))

    # Set column widths for readability
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 30


# =============================================================================
# Main
# =============================================================================

def main():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    build_tab1_constants(wb)
    build_tab2_relevance(wb)
    build_tab3_semantic_filter(wb)
    build_tab4_iterative_answer(wb)
    build_tab5_deep_research(wb)
    build_tab6_pipeline(wb)
    build_tab7_eval_api(wb)
    build_tab8_not_implemented(wb)

    output_path = "/home/user/kh/docs/kh_scoring_math.xlsx"
    wb.save(output_path)
    print(f"Saved: {output_path}")
    print(f"Tabs: {wb.sheetnames}")


if __name__ == "__main__":
    main()
